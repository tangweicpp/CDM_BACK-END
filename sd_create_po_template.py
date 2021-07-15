import os
import uuid
import conn_db as conn
from openpyxl import load_workbook
from xlrd import open_workbook


def xstr(s):
    return '' if s is None else str(s).strip()


MAX_COL = 50
MAX_ROW = 50
COL_NAME_COLOR = 'FFFFFF00'
CELL_COLOR = 'FF7030A0'

WEB_STATIC_FILE_DIR = '../cdm1.1_web/static/file/'
# WEB_STATIC_FILE_DIR_TEST = '/www/wwwroot/lps/static/file/'
WEB_STATIC_FILE_DIR_TEST = '/www/wwwroot/cmp/static/file/'
WEB_STATIC_PIC_DIR = '../cdm1.1_web/static/pic/'
# WEB_STATIC_PIC_DIR_TEST = '/www/wwwroot/lps/static/pic/'
WEB_STATIC_PIC_DIR_TEST = '/www/wwwroot/cmp/static/pic/'


def upload_po_template(po_original_file, po_compare_file, po_pic_file, po_header):
    file_dirs = save_file(
        po_original_file, po_compare_file, po_pic_file, po_header)

    keys_dict = get_config(file_dirs['compare'], po_header)
    if not keys_dict:
        print("无法获取配置项")
        return False

    update_config(file_dirs['orig'], po_header, keys_dict)
    save_config(file_dirs, po_header, keys_dict)


# 保存PO文件
def save_file(po_original_file, po_compare_file, po_pic_file, po_header):
    file_dirs = {}
    # 保存原始文件
    local_dir = os.path.join(os.getcwd(), WEB_STATIC_FILE_DIR_TEST +
                             po_header['cust_code'])

    if not os.path.exists(local_dir):
        os.makedirs(local_dir)
    file_path = os.path.join(
        local_dir, get_rand_id(6) + po_original_file.filename)
    po_original_file.save(file_path)
    file_dirs['orig'] = file_path

    # 保存对照文件
    local_dir = os.path.join(os.getcwd(), WEB_STATIC_FILE_DIR_TEST +
                             po_header['cust_code'])
    if not os.path.exists(local_dir):
        os.makedirs(local_dir)
    file_path = os.path.join(
        local_dir, get_rand_id(6)+po_compare_file.filename)
    po_compare_file.save(file_path)
    file_dirs['compare'] = file_path

    # 保存截图文件
    local_dir = os.path.join(os.getcwd(), WEB_STATIC_PIC_DIR_TEST +
                             po_header['cust_code'])
    if not os.path.exists(local_dir):
        os.makedirs(local_dir)
    file_path = os.path.join(local_dir, get_rand_id(6)+po_pic_file.filename)
    po_pic_file.save(file_path)
    file_dirs['pic'] = file_path

    return file_dirs


# 解析对照文件
def get_config(file_dir, po_header):
    keys_dict = {}
    wb = load_workbook(file_dir)
    ws = wb.get_sheet_by_name(wb.sheetnames[0])

    for row in range(1, MAX_ROW):
        for col in range(1, MAX_COL):
            cell_val = ws.cell(col, row).value

            if cell_val:
                # 背景颜色分辩
                cell_color = ws.cell(col, row).fill.fgColor.rgb

                # By col_name
                if cell_color == COL_NAME_COLOR:
                    if '/' in cell_val:
                        str_array = cell_val.split('/')
                        for str_s in str_array:
                            keys_dict[str_s] = {}
                            keys_dict[str_s]['col_row'] = {
                                'col': col, 'row': row}
                            keys_dict[str_s]['col_name'] = ''
                    else:
                        keys_dict[cell_val] = {}
                        keys_dict[cell_val]['col_row'] = {
                            'col': col, 'row': row}
                        keys_dict[cell_val]['col_name'] = ''

                # By col_row
                if cell_color == CELL_COLOR:
                    keys_dict[cell_val] = {}
                    keys_dict[cell_val]['col_row'] = {
                        'col': col, 'row': row}

    # waferid关键因子检查
    # if not 'lot_id' in keys_dict:
    #     return None

    return keys_dict


# 更新配置
def update_config(file_dir, po_header, keys_dict):
    file_suffix = os.path.splitext(file_dir)[-1].upper()
    if file_suffix == '.XLSX':
        wb = load_workbook(file_dir)
        ws = wb.get_sheet_by_name(wb.sheetnames[0])
        for key, val in keys_dict.items():
            if 'col_name' in val:
                val['col_name'] = xstr(ws.cell(
                    val['col_row']['col'], val['col_row']['row']).value)

    elif file_suffix == '.XLS':
        ws = open_workbook(file_dir).sheets()[0]
        for key, val in keys_dict.items():
            if 'col_name' in val:
                val['col_name'] = xstr(ws.cell_value(
                    val['col_row']['col']-1, val['col_row']['row']-1))


# 保存配置
def save_config(file_dirs, po_header, keys_dict):
    con = conn.HanaConn()
    parse_by_col_name = '0'
    parse_by_col_row = '0'

    # 插入模板总表
    po_pic_path = './'+file_dirs['pic'][file_dirs['pic'].index('static/pic'):]
    po_file_path = './' + \
        file_dirs['orig'][file_dirs['orig'].index('static/file'):]
    po_file_name = os.path.split(po_file_path)[1]

    sql = f"SELECT TEMPLATE_SN FROM ZM_CDM_PO_TEMPLATE_LIST WHERE CUST_CODE = '{po_header['cust_code']}' AND FILE_LEVEL ='{po_header['template_type']}' AND FILE_DESC='{po_header['template_name']}' "
    results = con.query(sql)
    if results:
        temp_sn = xstr(results[0][0])
        # 插入模板明细表
        key_fields_list = ['po_id', 'po_date', 'po_qty', 'cust_code', 'wafer_qty',
                           'customer_device', 'fab_device', 'ht_pn', 'process', 'code', 'product_pn', 'lot_id', 'wafer_id', 'wafer_id_str', 'grossbin_count', 'passbin_count', 'failbin_count', 'mark_code',
                           'add_0', 'add_1', 'add_2', 'add_3', 'add_4', 'add_5', 'add_6', 'add_7', 'add_8', 'add_9', 'add_10', 'add_11', 'add_12', 'add_13', 'add_14', 'add_15', 'add_16', 'add_17',
                           'add_18', 'add_19', 'add_20', 'add_21', 'add_22', 'add_23', 'add_24', 'add_25', 'add_26', 'add_27', 'add_28', 'add_29', 'add_30',
                           'wafer_pcs_price', 'wafer_die_price', 'address_code']

        for key in key_fields_list:
            key_obj = keys_dict.get(key)
            if key_obj:
                col_name = key_obj.get('col_name')
                col_row = key_obj.get('col_row')
                col_row = str(col_row['col']) + ':' + str(col_row['row'])

                if con.query(f"SELECT * FROM ZM_CDM_PO_TEMPLATE_ITEM zcpti WHERE id = '{temp_sn}' AND FIELD_NAME = '{key}' "):
                    if col_name:
                        parse_by_col_name = '1'
                        sql = f'''UPDATE ZM_CDM_PO_TEMPLATE_ITEM SET PARSE_METHOD='1',PARSE_SITE='{col_name}',PARSE_SITE_COL_ROW='{col_row}',FLAG='1' WHERE id = '{temp_sn}' AND FIELD_NAME = '{key}'  '''
                        con.exec_c(sql)
                    else:
                        parse_by_col_row = '1'
                        sql = f'''UPDATE ZM_CDM_PO_TEMPLATE_ITEM SET PARSE_METHOD='2',PARSE_SITE='{col_row}',PARSE_SITE_COL_ROW='{col_row}',FLAG='1' WHERE id = '{temp_sn}' AND FIELD_NAME = '{key}' '''
                        con.exec_c(sql)
                else:
                    if col_name:
                        parse_by_col_name = '1'
                        sql = f'''insert into ZM_CDM_PO_TEMPLATE_ITEM(ID,FIELD_NAME,FIELD_DESC,FIELD_TYPE,PARSE_METHOD,PARSE_SITE,PARSE_SITE_COL_ROW,
                        IGNORE_CHAR_1,IGNORE_CHAR_2,IGNORE_CHAR_3,START_CHAR,END_CHAR,FRONT_FIELD,BEHIND_FIELD,MES_FIELD_NAME,FLAG)
                        values('{temp_sn}','{key}','','','1','{col_name}','{col_row}','','','','','','','','','1')
                        '''
                        con.exec_c(sql)
                    else:
                        parse_by_col_row = '1'
                        sql = f'''insert into ZM_CDM_PO_TEMPLATE_ITEM(ID,FIELD_NAME,FIELD_DESC,FIELD_TYPE,PARSE_METHOD,PARSE_SITE,PARSE_SITE_COL_ROW,
                        IGNORE_CHAR_1,IGNORE_CHAR_2,IGNORE_CHAR_3,START_CHAR,END_CHAR,FRONT_FIELD,BEHIND_FIELD,MES_FIELD_NAME,FLAG)
                        values('{temp_sn}','{key}','','','2','{col_row}','{col_row}','','','','','','','','','1')
                        '''
                        con.exec_c(sql)

    else:
        temp_sn = get_rand_id(6)

        # 删除前一个配置
        sql = f"delete from ZM_CDM_PO_TEMPLATE_LIST where cust_code= '{po_header['cust_code']}' and FILE_LEVEL = '{po_header['template_type']}' and FILE_DESC = '{po_header['template_name']}' "
        con.exec_c(sql)

        # sql = f"delete from ZM_CDM_LABEL_VALUE_LOOKUP where TYPE='CUST_CODE' and value = '{po_header['cust_code']}'  "
        # con.exec_c(sql)

        # 插入新配置
        sql = f'''insert into ZM_CDM_PO_TEMPLATE_LIST(CUST_CODE,TEMPLATE_FILE,TEMPLATE_PIC,KEY_LIST,FILE_LEVEL,FLAG,ACCEPT,
        FILE_DESC,FILE_URL,CREATE_DATE,CREATE_BY,TEMPLATE_SN)
        values('{po_header['cust_code']}','{po_file_name}','{po_pic_path}','LOT|WAFER|DIES|CUSTPN|FABPN|PO','{po_header['template_type']}','1',
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/vnd.ms-excel',
        '{po_header['template_name']}','{po_file_path}',NOW(),'{po_header['user_name']}','{temp_sn}')
        '''

        print(sql)
        con.exec_c(sql)

        # 判断是否已存在
        sql = f"SELECT * FROM ZM_CDM_LABEL_VALUE_LOOKUP WHERE TYPE='CUST_CODE' and value = '{po_header['cust_code']}' "
        results = con.query(sql)
        if not results:
            sql = f"insert into ZM_CDM_LABEL_VALUE_LOOKUP(TYPE,LABEL,VALUE,DESCRIPE,ENABLE,ID) values('CUST_CODE','{po_header['cust_code']}','{po_header['cust_code']}','旧客户代码','Y',ZM_CDM_LABEL_VALUE_LOOKUP_SEQ.NEXTVAL)"
            con.exec_c(sql)

        # 插入模板明细表
        key_fields_list = ['po_id', 'po_date', 'po_qty', 'cust_code', 'wafer_qty',
                           'customer_device', 'fab_device', 'ht_pn', 'process', 'code', 'product_pn', 'lot_id', 'wafer_id', 'wafer_id_str', 'grossbin_count', 'passbin_count', 'failbin_count', 'mark_code',
                           'add_0', 'add_1', 'add_2', 'add_3', 'add_4', 'add_5', 'add_6', 'add_7', 'add_8', 'add_9', 'add_10', 'add_11', 'add_12', 'add_13', 'add_14', 'add_15', 'add_16', 'add_17',
                           'add_18', 'add_19', 'add_20', 'add_21', 'add_22', 'add_23', 'add_24', 'add_25', 'add_26', 'add_27', 'add_28', 'add_29', 'add_30',
                           'wafer_pcs_price', 'wafer_die_price', 'address_code']

        for key in key_fields_list:
            key_obj = keys_dict.get(key)
            if key_obj:
                col_name = key_obj.get('col_name')
                col_row = key_obj.get('col_row')
                col_row = str(col_row['col']) + ':' + str(col_row['row'])
                if col_name:
                    parse_by_col_name = '1'
                    sql = f'''insert into ZM_CDM_PO_TEMPLATE_ITEM(ID,FIELD_NAME,FIELD_DESC,FIELD_TYPE,PARSE_METHOD,PARSE_SITE,PARSE_SITE_COL_ROW,
                    IGNORE_CHAR_1,IGNORE_CHAR_2,IGNORE_CHAR_3,START_CHAR,END_CHAR,FRONT_FIELD,BEHIND_FIELD,MES_FIELD_NAME,FLAG)
                    values('{temp_sn}','{key}','','','1','{col_name}','{col_row}','','','','','','','','','1')
                    '''
                    con.exec_c(sql)
                else:
                    parse_by_col_row = '1'
                    sql = f'''insert into ZM_CDM_PO_TEMPLATE_ITEM(ID,FIELD_NAME,FIELD_DESC,FIELD_TYPE,PARSE_METHOD,PARSE_SITE,PARSE_SITE_COL_ROW,
                    IGNORE_CHAR_1,IGNORE_CHAR_2,IGNORE_CHAR_3,START_CHAR,END_CHAR,FRONT_FIELD,BEHIND_FIELD,MES_FIELD_NAME,FLAG)
                    values('{temp_sn}','{key}','','','2','{col_row}','{col_row}','','','','','','','','','1')
                    '''
                    con.exec_c(sql)
            # else:
            #     sql = f'''insert into ZM_CDM_PO_TEMPLATE_ITEM(id,FIELD_NAME,FLAG) values('{temp_sn}','{key}','0') '''
            #     con.exec_c(sql)

        # 插入模板头表
        if keys_dict.get('lot_id'):
            sheet_header = keys_dict['lot_id']['col_row']['col'] - 1

        elif keys_dict.get('customer_device'):
            sheet_header = keys_dict['customer_device']['col_row']['col'] - 1

        else:
            sheet_header = 0

        sql = f'''insert into ZM_CDM_PO_TEMPLATE_HEADER(ID,SHEET_INDEX,SHEET_HEADER,SHEET_MAX_COLS,SHEET_MAX_ROWS,PARSE_BY_COL_ROW,PARSE_BY_COL_NAME)
        values('{temp_sn}',0,{sheet_header},80,30,'{parse_by_col_row}','{parse_by_col_name}')
        '''

        print(sql)
        con.exec_c(sql)


def get_rand_id(id_len):
    return str(uuid.uuid1())[:id_len]
