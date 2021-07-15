import conn_db as conn
import os
from openpyxl import load_workbook


def xstr(s):
    return '' if s is None else str(s).strip()


# 渲染SQL数据
def renderData(items_data, header_data, file_name):
    # 加载模板xl
    wb = load_workbook('docTemplates/template_02.xlsx')
    ws = wb.get_sheet_by_name(wb.sheetnames[0])

    cols = len(header_data)
    rows = len(items_data)

    for i in range(1, cols+2):
        if i == 1:
            ws.cell(column=i, row=1, value="No.")
        else:
            ws.cell(column=i, row=1, value=header_data[i-2][0])

    for i in range(1, rows+1):
        for j in range(1, cols+2):
            if j == 1:
                cell_val = i
            else:
                cell_val = xstr(items_data[i-1][j-2])

            ws.cell(column=j, row=i+1, value=cell_val)

    # 文件名
    con = conn.HanaConn()
    sql = f"select ZM_CDM_FILE_ID_SEQ.NEXTVAL from dummy"
    seq = xstr(con.query(sql)[0][0])

    if file_name:
        xl_part_name = seq + "_" + file_name
    else:
        xl_part_name = seq + "_" + "已上传订单WO数据_.xlsx"

    if file_name:
        xl_part_name = file_name
    else:
        xl_part_name = "已上传订单WO数据_.xlsx"

    if not os.path.exists("/root/docs/"):
        os.makedirs("/root/docs/")

    xl_name = '/root/docs/' + xl_part_name

    # 保存
    wb.save(xl_name)

    # 关闭
    wb.close()

    print(xl_name)
    # 保存文件记录
    sql = f"""
        INSERT INTO ZM_CDM_FILE_ID_LIST(FILE_PATH,FILE_NAME,FILE_ABS_PATH,FILE_ID,CREATE_DATE,FLAG,REMARK)   
        values( '/root/docs/','{xl_part_name}','{xl_name}','{seq}',NOW(),'Y','')
    """
    con.exec_c(sql)

    return seq


# 渲染基础数据
def renderData2(items_data, header_data, file_name):
    # 加载模板xl
    wb = load_workbook('docTemplates/template_02.xlsx')
    ws = wb.get_sheet_by_name(wb.sheetnames[0])

    cols = len(header_data)
    rows = len(items_data)

    for i in range(1, cols+2):
        if i == 1:
            ws.cell(column=i, row=1, value="No.")
        else:
            ws.cell(column=i, row=1, value=header_data[i-2])

    for i in range(1, rows+1):
        for j in range(1, cols+2):
            if j == 1:
                cell_val = i
            else:
                cell_val = xstr(items_data[i-1][j-2])

            ws.cell(column=j, row=i+1, value=cell_val)

    # 文件名
    con = conn.HanaConn()
    sql = f"select ZM_CDM_FILE_ID_SEQ.NEXTVAL from dummy"
    seq = xstr(con.query(sql)[0][0])

    if file_name:
        xl_part_name = seq + "_" + file_name
    else:
        xl_part_name = seq + "_" + "已上传订单WO数据_.xlsx"

    if file_name:
        xl_part_name = file_name
    else:
        xl_part_name = "已上传订单WO数据_.xlsx"

    if not os.path.exists("/root/docs/"):
        os.makedirs("/root/docs/")

    xl_name = '/root/docs/' + xl_part_name

    # 保存
    wb.save(xl_name)

    # 关闭
    wb.close()

    print(xl_name)
    # 保存文件记录
    sql = f"""
        INSERT INTO ZM_CDM_FILE_ID_LIST(FILE_PATH,FILE_NAME,FILE_ABS_PATH,FILE_ID,CREATE_DATE,FLAG,REMARK)   
        values( '/root/docs/','{xl_part_name}','{xl_name}','{seq}',NOW(),'Y','')
    """
    con.exec_c(sql)

    return seq



# 开始执行
def trans_sql(sql, file_name=""):
    con = conn.HanaConn()
    items_data, header_data = con.query2(sql)
    if not items_data:
        return False

    # 渲染数据
    file_id = renderData(items_data, header_data, file_name)
    return file_id


# 开始执行DW
def trans_sql_dw(sql, file_name=""):
    con = conn.HanaConnDW()
    items_data, header_data = con.query2(sql)
    if not items_data:
        return False

    # 渲染数据
    file_id = renderData(items_data, header_data, file_name)
    return file_id


if __name__ == "__main__":
    sql = f"""SELECT CUST_CODE as "客户代码",SAP_CUST_CODE as "SAP客户代码",TRAD_CUST_CODE as "交易客户" FROM ZM_CDM_PO_ITEM  WHERE MO_ID = 'ATC-210213205'  """


    trans_sql(sql, "测试.xlsx")
